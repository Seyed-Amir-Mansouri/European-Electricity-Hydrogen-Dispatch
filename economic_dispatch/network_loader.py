"""Transport lines + global price scalars, from the networks parquet database.

``build_networks_db`` converts ``Networks.xlsx`` to ``networks_2030.parquet``
(all lines of both carriers, losses already resolved, plus the CO2/gas prices);
``load_networks`` reads that database and filters the lines to the selected
zones. At runtime only the parquet is needed — not the Excel file.

Networks.xlsx layout: each network sheet has two side-by-side blocks —
  cols A-D : From, To, Length (km), Loss Fraction   -> distance/loss per pair
  cols F-I : From, To, From-To Capacity, To-From Cap -> directional MW limits
The two blocks cover different pair sets, so the capacity block is the
authoritative line list and losses are looked up by unordered {From, To} pair.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd

from .config import DEFAULT_DATA_DIR, DEFAULT_NETWORKS_DB, DEFAULT_H2_REF

SHEET_ELEC = "Electricity Lines"
SHEET_H2 = "Hydrogen Pipelines"
SHEET_DATA = "Data"
_CARRIERS = [("electricity", SHEET_ELEC), ("hydrogen", SHEET_H2)]
_CO2, _GAS = "CO2 Price (EUR/ton)", "Gas Price (EUR/MWh)"
# Reference-grid interconnector-hub codes -> the country they represent.
_H2_HUB_COUNTRY = {"IBIT": "IT", "IBFI": "FI"}


@dataclass
class Line:
    frm: str
    to: str
    cap_ft: float   # MW capacity in the from->to direction
    cap_tf: float   # MW capacity in the to->from direction
    loss: float     # fractional loss on the line (0-1)


@dataclass
class NetworkData:
    elec: list[Line]
    hydrogen: list[Line]
    co2_price: float      # EUR/ton
    gas_price: float      # EUR/MWh


# --------------------------------------------------------------------------- #
# Build:  Networks.xlsx  ->  networks_2030.parquet
# --------------------------------------------------------------------------- #
def _loss_lookup(ws) -> dict[frozenset, float]:
    """Build {frozenset({From, To}) -> loss fraction} from the left block."""
    out: dict[frozenset, float] = {}
    for r in range(2, ws.max_row + 1):
        frm, to, _length, loss = (ws.cell(r, c).value for c in (1, 2, 3, 4))
        if isinstance(frm, str) and isinstance(to, str):
            try:
                out[frozenset({frm, to})] = float(loss) if loss is not None else 0.0
            except (TypeError, ValueError):
                out[frozenset({frm, to})] = 0.0
    return out


def _read_lines_all(ws) -> list[tuple]:
    """Every capacity-block line (frm, to, cap_ft, cap_tf, loss); no zone filter."""
    losses = _loss_lookup(ws)
    rows = []
    for r in range(2, ws.max_row + 1):
        frm, to, cap_ft, cap_tf = (ws.cell(r, c).value for c in (6, 7, 8, 9))
        if not (isinstance(frm, str) and isinstance(to, str)):
            continue
        try:
            ft = float(cap_ft) if cap_ft is not None else 0.0
            tf = float(cap_tf) if cap_tf is not None else 0.0
        except (TypeError, ValueError):
            ft = tf = 0.0
        rows.append((frm, to, ft, tf, losses.get(frozenset({frm, to}), 0.0)))
    return rows


def _h2_reference_caps(ref_path: Path) -> dict[tuple, float]:
    """Directional H2 border capacities (GW) from ReferenceGrid_Hydrogen '2030'.

    Returns {(countryA, countryB): GW A->B}. Border header ``A-B`` gives dir1 =
    A->B and dir2 = B->A; interconnector-hub codes are mapped to their country.
    """
    import openpyxl
    ref_path = Path(ref_path)
    if not ref_path.exists():
        return {}
    wb = openpyxl.load_workbook(ref_path, read_only=True, data_only=True)
    cap: dict[tuple, float] = {}
    for border, d1, d2 in ((r[0], r[1], r[2]) for r in wb["2030"].iter_rows(min_row=2, values_only=True)):
        if not isinstance(border, str) or "-" not in border:
            continue
        a, b = border.split("-", 1)
        a, b = _H2_HUB_COUNTRY.get(a, a), _H2_HUB_COUNTRY.get(b, b)
        try:
            cap[(a, b)], cap[(b, a)] = float(d1), float(d2)
        except (TypeError, ValueError):
            pass
    wb.close()
    return cap


def build_networks_db(data_dir: Path = DEFAULT_DATA_DIR, out: Path = DEFAULT_NETWORKS_DB,
                      h2_ref: Path = DEFAULT_H2_REF) -> Path:
    """Convert Networks.xlsx to networks_2030.parquet (all lines + prices).

    Hydrogen line capacities are overridden by ReferenceGrid_Hydrogen (GW x 1000
    -> MW, direction-aligned) wherever the line's country pair has a reference
    border; other lines keep their Networks.xlsx capacities.
    """
    import openpyxl
    path = Path(data_dir) / "Networks.xlsx"
    if not path.exists():
        raise FileNotFoundError(f"Networks workbook not found: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ref = _h2_reference_caps(h2_ref)
    rows, overridden = [], 0
    for carrier, sheet in _CARRIERS:
        for frm, to, ft, tf, loss in _read_lines_all(wb[sheet]):
            if carrier == "hydrogen" and (frm[:2], to[:2]) in ref:
                ft = ref[(frm[:2], to[:2])] * 1000.0            # GW -> MW
                tf = ref.get((to[:2], frm[:2]), 0.0) * 1000.0
                overridden += 1
            rows.append(dict(carrier=carrier, frm=frm, to=to,
                             cap_from_to_mw=ft, cap_to_from_mw=tf, loss_fraction=loss))
    ws = wb[SHEET_DATA]
    for name, col in [(_CO2, 1), (_GAS, 2)]:
        rows.append(dict(carrier="prices", frm=name, to=None,
                         cap_from_to_mw=float(ws.cell(2, col).value or 0.0),
                         cap_to_from_mw=None, loss_fraction=None))
    wb.close()
    if ref:
        print(f"  applied ReferenceGrid_Hydrogen to {overridden} H2 lines (GW x 1000)")
    out = Path(out)
    out.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(rows).to_parquet(out, index=False)
    return out


# --------------------------------------------------------------------------- #
# Load:  networks_2030.parquet  ->  NetworkData (filtered to selected zones)
# --------------------------------------------------------------------------- #
def load_networks(zones: list[str], db_path: Path = DEFAULT_NETWORKS_DB) -> NetworkData:
    db_path = Path(db_path)
    if not db_path.exists():
        raise FileNotFoundError(
            f"Networks database not found: {db_path}. Build it with `python build_db.py`.")
    df = pd.read_parquet(db_path)
    zset = set(zones)

    def lines_for(carrier: str) -> list[Line]:
        sub = df[df["carrier"] == carrier]
        out: list[Line] = []
        for frm, to, ft, tf, loss in zip(sub["frm"], sub["to"], sub["cap_from_to_mw"],
                                          sub["cap_to_from_mw"], sub["loss_fraction"]):
            if frm in zset and to in zset and frm != to:  # internal lines only
                out.append(Line(frm, to, float(ft), float(tf), float(loss)))
        return out

    prices = df[df["carrier"] == "prices"].set_index("frm")["cap_from_to_mw"].to_dict()
    return NetworkData(lines_for("electricity"), lines_for("hydrogen"),
                       float(prices.get(_CO2, 0.0)), float(prices.get(_GAS, 0.0)))
